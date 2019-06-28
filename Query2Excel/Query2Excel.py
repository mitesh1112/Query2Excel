import os
import datetime
import argparse
import xml.etree.ElementTree as ET
import openpyxl
import pyodbc


def generate_excel(sheet_name, cnn_str, query, work_book):
    """
    Generates the Excel sheet for the sheet passed.
    :param sheet_name:
    :param cnn_str:
    :param query:
    :param work_book:
    :return:
    """

    print(f'Preparing the sheet {sheet_name}')

    sheet = work_book.create_sheet(title=sheet_name)

    cnn = pyodbc.connect(cnn_str)
    cur = cnn.cursor()
    cur.execute(query)
    rows = cur.fetchall()

    if not rows:
        return

    row_no = 1

    # Add Header
    for i in range(len(rows[0].cursor_description)):
        cell = sheet.cell(row=row_no, column=i + 1)
        # First item in cursor_description is the field name
        cell.value = rows[0].cursor_description[i][0]

        if isinstance(rows[0][i], datetime.datetime):
            cell.number_format = 'dd-MMM-yy'
            cell.alignment = openpyxl.styles.Alignment(horizontal='center')

    row_no += 1

    for row in rows:
        for j in range(len(row)):
            cell = sheet.cell(row=row_no, column=j + 1)
            cell.value = row[j]
            if isinstance(row[j], datetime.datetime):
                cell.number_format = 'dd-MMM-yy'
                cell.alignment = openpyxl.styles.Alignment(horizontal='center')

        row_no += 1


def process_xml(xml_file_path):
    """
    Opens the XML file and generates Excel Worksheets in Excel workbook for each query nodes.
    :param xml_file_path: XML file to be read to generate the Excel workbook.
    :return: None
    """

    tree = ET.parse(xml_file_path)
    root = tree.getroot()

    # Connection string given in queries root node
    global_cnn_str = ''
    if 'connection' in root.attrib:
        global_cnn_str = root.attrib['connection']


    excel_file_path = os.path.join(root.attrib['filepath'], root.attrib['filename'])
    if 'appenddate' in root.attrib:
        if root.attrib['appenddate'] == 'yes':
            today = datetime.date.today()
            excel_file_path = f"{excel_file_path}_{today.strftime('%d-%b-%Y')}"
    excel_file_path = excel_file_path + '.xlsx'

    work_book = openpyxl.Workbook()
    for query in root:
        sheet_name = query.attrib['name']

        # If connection attribute provided at query, pick connection string from the query node.
        if 'connection' in query.attrib:
            cnn_str = query.attrib['connection']
        else:
            cnn_str = global_cnn_str

        sql = query.find('sql')
        # If file attribute given..
        if 'file' in sql.attrib:
            # Pick query from the sql file
            with open(sql.attrib['file'], 'r') as sqlfile:
                sql_text = sqlfile.read()
        else:
            sql_text = sql.text

        generate_excel(sheet_name, cnn_str, sql_text, work_book)

    work_book.save(excel_file_path)
    print(f'File {excel_file_path} created successfully.')


def main():
    """
    The main function when this module is called..
    :return: None
    """

    aparser = argparse.ArgumentParser(
        description='Mitesh: Query 2 Excel.  Give a XML file with query and have output in Excel .',
        epilog='You are seeing EPILOG'
    )
    aparser.add_argument('xml', type=str, help='xml file having queries, connections, etc...')
    args = aparser.parse_args()
    xml_file_path = args.xml
    process_xml(xml_file_path)


if __name__ == '__main__':
    main()
